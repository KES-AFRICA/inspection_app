# generate_journal.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi des Évolutions Majeures"
    
    # Activer le quadrillage à l'affichage
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "N°", "Nom du chantier", "Période / Date", "Objectif du chantier",
        "Grandes modifications apportées", "Bénéfices pour l'application", "Statut d'avancement"
    ]
    
    chantiers_data = [
        [
            "1",
            "Restructuration de l'application (Clean Architecture)",
            "08/07/2026 - 09/07/2026",
            "Découpler la logique métier de la base de données de persistance locale Hive et de l'interface utilisateur pour rendre l'application robuste, évolutive et testable.",
            "Migration de l'intégralité des 10 modules de la séquence d'inspection vers un modèle à 3 couches (Domain pour la logique pure et Use Cases, Data pour l'infrastructure Hive et Presentation pour l'interface utilisateur). Injection des dépendances gérée via GetIt.",
            "1. Robustesse et modularité accrues du code.\n2. Indépendance totale vis-à-vis du moteur de base de données Hive.\n3. Possibilité d'écrire des tests automatisés sur la logique métier.",
            "Terminé"
        ],
        [
            "2",
            "Introduction du State Management avec Riverpod",
            "09/07/2026",
            "Remplacer l'état local éparpillé (setState) par une gestion d'état réactive, centralisée et robuste, garantissant zéro perte de données pour les inspecteurs sur le terrain.",
            "Mise en place de ProviderScope racine, création de providers immutables de Use Cases, et migration progressive de toutes les étapes de la séquence d'inspection (étapes 'Renseignements Généraux', 'JSA', 'Documents requis' et 'Schéma des installations' migrées, autres en cours).",
            "1. Cycle de vie de l'état applicatif unifié et réactif.\n2. Sauvegarde asynchrone transparente en tâche de fond pour l'inspecteur.\n3. Sécurisation et fluidité accrue de la saisie utilisateur sans frame lags.",
            "En cours"
        ]
    ]
    
    # Écriture des en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Bleu foncé corporate
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='medium', color='111111'),
            bottom=Side(style='medium', color='111111')
        )
    
    # Écriture des données
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for row_idx, row_data in enumerate(chantiers_data, 2):
        row_fill = PatternFill(start_color="F7FAFC" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Segoe UI", size=10)
            cell.fill = row_fill
            cell.border = thin_border
            
            # Alignements spécifiques
            if col_idx in [1, 3, 7]:  # N°, Date, Statut
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
            # Style du statut d'avancement
            if col_idx == 7:
                if value == "Terminé":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="1E3A8A") # Bleu foncé d'achèvement
                    cell.fill = PatternFill(start_color="DBEAFE", fill_type="solid") # Fond bleu clair
                elif value == "En cours":
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color="B45309") # Orange/marron
                    cell.fill = PatternFill(start_color="FEF3C7", fill_type="solid") # Fond jaune clair

    # Hauteurs de lignes
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 95   # Chantier 1
    ws.row_dimensions[3].height = 95   # Chantier 2
    
    # Largeurs de colonnes
    col_widths = {
        1: 6,   # N°
        2: 30,  # Nom du chantier
        3: 20,  # Période
        4: 38,  # Objectif
        5: 45,  # Modifications
        6: 38,  # Bénéfices
        7: 18   # Statut
    }
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Sauvegarder
    output_path = "/home/andelson-teufack/Bureau/Projets/KES/inspection_app/Documentation des corrections sur InspectApp.xlsx"
    wb.save(output_path)
    print(f"✅ Document Excel de suivi mis à jour avec succès : {output_path}")

if __name__ == "__main__":
    generate_excel()
